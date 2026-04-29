"""Spreadsheet tools backed by openpyxl. Handy for coding tasks that involve
.xlsx data dumps without firing up Excel.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .registry import BuiltinTool


_MAX_ROWS = 200
_MAX_COLS = 40


def _read_sheet(args: dict[str, Any]) -> str:
    path = Path(args.get("path", ""))
    sheet = args.get("sheet")
    max_rows = int(args.get("max_rows", _MAX_ROWS))
    max_cols = int(args.get("max_cols", _MAX_COLS))

    if not path.exists():
        return f"File not found: {path}"
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return f"Failed to open {path}: {e}"

    ws = wb[sheet] if sheet else wb.active
    rows: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            rows.append([f"[... truncated at {max_rows} rows ...]"])
            break
        rows.append(list(row[:max_cols]))

    # ReadOnlyWorksheet doesn't expose .dimensions reliably; compute a safe shape.
    info = {
        "path": str(path),
        "sheet": ws.title,
        "sheet_names": wb.sheetnames,
        "rows_returned": len(rows),
        "cols_max": max((len(r) for r in rows), default=0),
        "rows": rows,
    }
    wb.close()
    return json.dumps(info, indent=2, default=str)


def _list_sheets(args: dict[str, Any]) -> str:
    path = Path(args.get("path", ""))
    if not path.exists():
        return f"File not found: {path}"
    try:
        wb = load_workbook(path, read_only=True)
    except Exception as e:
        return f"Failed to open {path}: {e}"
    names = wb.sheetnames
    wb.close()
    return "\n".join(names) if names else "(no sheets)"


def _write_sheet(args: dict[str, Any]) -> str:
    path = Path(args.get("path", ""))
    sheet = args.get("sheet", "Sheet1")
    rows = args.get("rows", [])
    if not isinstance(rows, list):
        return "[error] 'rows' must be a list of lists"

    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
        if sheet in wb.sheetnames:
            del wb[sheet]
        ws = wb.create_sheet(sheet)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet

    for row in rows:
        ws.append(list(row) if isinstance(row, (list, tuple)) else [row])

    wb.save(path)
    wb.close()
    return f"Wrote {len(rows)} rows to {path} (sheet {sheet!r})"


TOOLS: list[BuiltinTool] = [
    BuiltinTool(
        name="excel_read",
        description="Read rows from an .xlsx sheet. Returns JSON with rows truncated for safety.",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string"},
                "sheet": {"type": "string", "description": "Sheet name (default: active)."},
                "max_rows": {"type": "integer", "default": _MAX_ROWS},
                "max_cols": {"type": "integer", "default": _MAX_COLS},
            },
            "required": ["path"],
        },
        handler=_read_sheet,
    ),
    BuiltinTool(
        name="excel_sheets",
        description="List sheet names in an .xlsx workbook.",
        input_schema={
            "type": "object",
            "properties": {"path": {"type": "string"}},
            "required": ["path"],
        },
        handler=_list_sheets,
    ),
    BuiltinTool(
        name="excel_write",
        description="Write a 2D list of rows to a sheet (creating the workbook if needed).",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string"},
                "sheet": {"type": "string", "default": "Sheet1"},
                "rows": {
                    "type": "array",
                    "description": "List of rows; each row is a list of cells.",
                    "items": {"type": "array"},
                },
            },
            "required": ["path", "rows"],
        },
        handler=_write_sheet,
    ),
]
